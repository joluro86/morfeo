import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Candidato, EquivalenciaTitulo, NovedadEquivalencia
from django.views import View
from django.core.files.storage import default_storage
from .models import EquivalenciaTitulo, ProgramaAcademicoSnies, NovedadEquivalencia
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.views import View
from django.db.models import Q
from .models import EquivalenciaTitulo, ProgramaAcademicoSnies, NovedadEquivalencia
from django.views import View
from django.shortcuts import render, get_object_or_404, redirect
from .models import EquivalenciaTitulo, NovedadEquivalencia, ProgramaAcademicoSnies
import openpyxl
from django.http import HttpResponse
from .models import EquivalenciaTitulo

def exportar_equivalencias_excel(request):
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equivalencias"

    # Encabezados
    encabezados = ['Doc Candidato', 'Vacante', 'Título', 'Otro Título', 'Nivel de Estudios', 'Equivalente SNIES']
    ws.append(encabezados)

    # Agregar datos
    equivalencias = EquivalenciaTitulo.objects.all()
    for e in equivalencias:
        ws.append([
            e.id_candidato,
            e.id_vacante,
            e.titulo,
            e.nivel_estudios,
            str(e.equivalente_snies) if e.equivalente_snies else 'Sin equivalente'
        ])

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equivalencias_titulo.xlsx'
    wb.save(response)
    return response
class CargarExcelEquivalenciaTitulo(View):
    def get(self, request):
        return render(request, 'equivalencia_titulo_cargar_excel.html')

    def post(self, request):
        excel = request.FILES.get('excel_file')
        if not excel:
            messages.error(request, 'Debes subir un archivo Excel.')
            return redirect('cargar_equivalencia_titulo')

        # Guardar temporalmente y abrir con openpyxl
        path = default_storage.save('tmp/' + excel.name, excel)
        wb = openpyxl.load_workbook(default_storage.path(path))
        ws = wb.active
        cont=1
        # Recorrer desde la segunda fila (sin encabezado)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[4]=="Bachiller" or row[4]=="Completar":
                continue
             
            if row[2]=="Busque su título académico en la lista" or row[2]=="" or row[2]==None or row[2]=='Busca tu título académico en la lista. Elige "No aplica" en caso de no encontrarlo' or row[2]=='No aplica (elige esta opción cuando no encuentres tu título)':
                titulo="No reporta"
            else:
                titulo=row[2]
            
            if row[3]=="Busque su título académico en la lista" or row[3]=="" or row[3]==None or row[3]=='Busca tu título académico en la lista. Elige "No aplica" en caso de no encontrarlo' or row[3]=='No aplica (elige esta opción cuando no encuentres tu título)':
                otro_titulo="No reporta"
            else:
                otro_titulo=row[3]   
            
            if titulo=="No reporta" and otro_titulo=="No reporta":
                continue
            
            if titulo!="No reporta":                    
                crear_noevdad_equivalencia(row[0],row[1],titulo,row[4],row[5],titulo + "-"+str(row[4]))   
            
            if otro_titulo!="No reporta":                    
                crear_noevdad_equivalencia(row[0],row[1],otro_titulo,row[4],row[5],otro_titulo + "-"+str(row[4]))                
                      
            cont+=1
        messages.success(request, 'Datos cargados correctamente desde el Excel.')
        return redirect('lista_equivalencia_titulo')


def crear_noevdad_equivalencia(id,idc,tit, ne, eqs, id_e):
    if not EquivalenciaTitulo.objects.filter(id_equivalencia=id_e).exists():
        EquivalenciaTitulo.objects.create(
                            id_vacante=id,            
                            id_candidato=idc,         
                            titulo=tit,               
                            nivel_estudios=ne,        
                            equivalente_snies=eqs,      
                            id_equivalencia=id_e 
                        )
class ListaEquivalenciaTitulo(View):
    def get(self, request):
        equivalencias = EquivalenciaTitulo.objects.all()
        return render(request, 'equivalencia_titulo_lista.html', {
            'equivalencias': equivalencias
        })
        
class ListaNovedades(View):
    def get(self, request):
        novedades = NovedadEquivalencia.objects.filter(descripcion="Sin titulo equivalente")
        return render(request, 'novedades_equivalencia_titulo.html', {'novedades': novedades})


class BuscarEquivalenciasView(View):
    def get(self, request):
        buscar_equivalencias_snies()
        messages.success(request, "Búsqueda de equivalencias completada.")
        return redirect('lista_novedades_equivalencia')

def limpiar_novedades_equivalencia(request):
    NovedadEquivalencia.objects.all().delete()
    EquivalenciaTitulo.objects.all().delete()
    return redirect('lista_novedades_equivalencia')
    
    
def buscar_equivalencias_snies():
    equivalencias = EquivalenciaTitulo.objects.all()
    cont1 = 1
    
    for eq in equivalencias:
        if eq.nivel_estudios=="Bachiller" or eq.nivel_estudios=="bachiller" or eq.nivel_estudios=="Completar" or eq.nivel_estudios=="completar":
            eq.equivalente_snies= "No aplica"
            eq.save()
            continue
        
        if eq.titulo is not None:
            pro1 = ProgramaAcademicoSnies.objects.filter(titulo_otorgado=eq.titulo).first()
            if pro1 is None:
                pro2= ProgramaAcademicoSnies.objects.filter(nombre_del_programa=eq.titulo).first()
            if pro1 is not None:              
                eq.equivalente_snies=pro1.titulo_otorgado
                eq.save()
            elif pro2 is not None:
                eq.equivalente_snies=pro2.nombre_del_programa
                eq.save()
            
            else: 
                # Aquí creamos la NovedadEquivalencia si no se encontró coincidencia en SNIES
                if eq.titulo=="No reporta":
                    eq.equivalente_snies= "No aplica"
                    eq.save()
                    continue
                    
                NovedadEquivalencia.objects.create(
                    descripcion="Sin titulo equivalente",
                    id_vacante=eq.id_vacante,
                    identificacion_candidato=eq.id_candidato,
                    titulo=eq.titulo,
                    nivel_estudios=eq.nivel_estudios,
                    equivalencia_snies=eq.id,
                )
                cont1 += 1 
             
        
class EditarNovedadesEquivalencia(View):
    def get(self, request, pk):
        equivalencia = get_object_or_404(EquivalenciaTitulo, pk=pk)
        lista_snies = ProgramaAcademicoSnies.objects.only('titulo_otorgado')  # más ligero
        return render(request, 'editar_novedades_equivalencia_titulo.html', {
            'equivalencia': equivalencia,
            'lista_snies': lista_snies
        })

    def post(self, request, pk):
        equivalencia = get_object_or_404(EquivalenciaTitulo, pk=pk)
        # Extraemos el dato enviado desde el input
        texto_sugerido = request.POST.get('nuevo_titulo_equivalente', '').strip()
        print("Texto sugerido:", texto_sugerido)
        # Buscamos en ProgramaAcademicoSnies un registro cuyo titulo_otorgado coincida con lo ingresado
        sugerido = ProgramaAcademicoSnies.objects.filter(titulo_otorgado=texto_sugerido).first()
        if sugerido:
            equivalencia.equivalente_snies = sugerido.titulo_otorgado
        else:
            equivalencia.equivalente_snies = texto_sugerido  # Se asigna el valor ingresado si no se encuentra en SNIES
        equivalencia.save()

        # Actualizamos la novedad asociada para quitar la marca de "Sin titulo equivalente"
        novedad = NovedadEquivalencia.objects.filter(equivalencia_snies=equivalencia.id).first()
        if novedad:
            if novedad.descripcion == "Sin titulo equivalente":
                novedad.descripcion = equivalencia.equivalente_snies
                novedad.save()

        messages.success(request, "Equivalencia actualizada correctamente.")
        return redirect('lista_equivalencia_titulo')

